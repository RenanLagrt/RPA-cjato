import os
import re
import io
import sys
import time
import psutil 
import subprocess
import pytesseract
import xml.etree.ElementTree as ET
import pandas as pd
import streamlit as st
from PIL import ImageOps, ImageEnhance
from pdf2image import convert_from_path
from datetime import datetime
from dotenv import load_dotenv
from selenium import webdriver
from openpyxl import load_workbook
from itertools import zip_longest
from datetime import datetime, timedelta
from openpyxl.cell.cell import MergedCell
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


poppler_path = os.path.join(os.path.expanduser("~"),"Downloads","Release-24.08.0-0","poppler-24.08.0","Library","bin")
tesseract_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = tesseract_path

class AutomaçãoSertras():

    def __init__(self, contratos, contrato_selecionado):
        load_dotenv()

        self.email = os.getenv("EMAIL")
        self.senha = os.getenv("SENHA")
        self.driver = None
        self.contratos = contratos
        self.contrato_selecionado = contrato_selecionado

    def get_info_contrato(self):
        doc_dp = self.contratos[self.contrato_selecionado]["documentos"]["DP"]
        doc_qsms = self.contratos[self.contrato_selecionado]["documentos"]["QSMS"]
        dir_dp = self.contratos[self.contrato_selecionado]["diretorio funcionarios"]["DP"]
        dir_qsms = self.contratos[self.contrato_selecionado]["diretorio funcionarios"]["QSMS"]
        mapeamento_para_documentos = self.contratos[self.contrato_selecionado]["mapeamentos"]["mapeamento_para_documentos"]
        mapeamento_para_datas = self.contratos[self.contrato_selecionado]["mapeamentos"]["mapeamento_para_datas"]
        mapeamento_para_comentarios = self.contratos[self.contrato_selecionado]["mapeamentos"]["mapeamento_para_comentarios"]
        xpath_botão_contrato = self.contratos[self.contrato_selecionado]["botão contrato"]

        return doc_dp, doc_qsms, dir_dp, dir_qsms, mapeamento_para_documentos, mapeamento_para_datas, mapeamento_para_comentarios, xpath_botão_contrato

    def initialize_driver(self):
        return webdriver.Chrome(service=Service(ChromeDriverManager().install())) 

    def login_sertras(self, xpath_botão_contrato):
        self.driver.get("https://gestaodeterceiros.sertras.com/escolha-um-contrato")
        self.driver.maximize_window()

        campo_email = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edtLoginInfo"]')))
        campo_email.send_keys(self.email)

        campo_senha = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edtLoginSenha"]')))
        campo_senha.send_keys(self.senha)

        botão_enter = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="btnLogin"]/div[2]')))
        botão_enter.click()

        botao_contrato = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath_botão_contrato)))
        botao_contrato.click()

        try:
            fechar_janela = WebDriverWait(self.driver, 2).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="notificationPopup"]/div/div/div[1]/button/span')))
            fechar_janela.click()
        
        except:
            pass

    def download_arquivo(self, tipo="pessoas"):
        mapeamento = {
            "pessoas": '//*[@id="sidebar-menu"]/div/ul/li[9]/ul/li[2]/ul/li[1]/a',
            "empresas": '//*[@id="sidebar-menu"]/div/ul/li[9]/ul/li[2]/ul/li[4]/a/span'
        }

        botão_relatório = WebDriverWait(self.driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/div/ul/li[9]/a/span[1]')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_relatório)
        botão_relatório.click()

        botão_integração = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/div/ul/li[9]/ul/li[2]/a/span[1]')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_integração)
        botão_integração.click()

        botão_final = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, mapeamento[tipo])))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_final)
        botão_final.click()

        marcar_todos = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="dashboard-v1"]/div[3]/div/div/div[2]/form/div[1]/div[1]/div/label/a[1]')))
        marcar_todos.click()

        botão_dowload = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="box-filter"]/button[3]')))
        botão_dowload.click()

    @staticmethod
    def wait_for_download(diretorio_downloads, timeout=45):
        arquivos_iniciais = set(os.listdir(diretorio_downloads))
        tempo_inicio = time.time()

        while time.time() - tempo_inicio < timeout:
            time.sleep(0.5) 
            arquivos_atuais = set(os.listdir(diretorio_downloads))
            novos_arquivos = arquivos_atuais - arquivos_iniciais

            arquivos_xls = [
                os.path.join(diretorio_downloads, arquivo)
                for arquivo in novos_arquivos
                if arquivo.lower().endswith(".xls") and os.path.isfile(os.path.join(diretorio_downloads, arquivo))
            ]

            if arquivos_xls:
                caminho_arquivo = max(arquivos_xls, key=os.path.getctime)
                tamanho_anterior = -1

                while time.time() - tempo_inicio < timeout:
                    time.sleep(0.5)
                    if os.path.exists(caminho_arquivo):
                        tamanho_atual = os.path.getsize(caminho_arquivo)
                        if tamanho_atual > 0 and tamanho_atual == tamanho_anterior:
                            return caminho_arquivo
                        tamanho_anterior = tamanho_atual

        raise TimeoutError("Erro: Tempo limite atingido para download.")

    @staticmethod
    def ler_xml(caminho_arquivo):
        try:
            tree = ET.parse(caminho_arquivo)
            root = tree.getroot()

        except ET.ParseError:
            raise ValueError("Erro ao processar XML. Verifique se o arquivo está correto.")

        namespace = '{urn:schemas-microsoft-com:office:spreadsheet}'
        dados, colunas = [], []
        is_header = True

        for row in root.iter(f'{namespace}Row'):
            linha = [
                cell[0].text.strip() if len(cell) > 0 and cell[0].tag == f'{namespace}Data' and cell[0].text else ''
                for cell in row
            ]
            if is_header:
                colunas = linha
                is_header = False
            else:
                if any(linha):  # Evita adicionar linhas vazias
                    dados.append(linha)

        return pd.DataFrame(dados, columns=colunas) 

    @staticmethod
    def tratar_tabela(tabela_sertras):
        colunas_para_remover = ["Contrato Terceiro", "Unidade", "Valor Preenchido", "Âmbito","Evento"]
        colunas_existentes = [col for col in colunas_para_remover if col in tabela_sertras.columns]

        if colunas_existentes:
            tabela_sertras = tabela_sertras.drop(columns=colunas_existentes)

        tabela_sertras = tabela_sertras.rename(columns={"Data da Última Análise": "Data Análise"})

        tabela_sertras = tabela_sertras.rename(columns=lambda x: x.upper())

        tabela_sertras["DOCUMENTO"] = tabela_sertras["DOCUMENTO"].replace({
                    "CTPS OU RELATÓRIO DO E-SOCIAL": "CTPS",
                    "DOCUMENTO DE IDENTIFICAÇÃO": "RG",
                    "FICHA DE REGISTRO": "FRE",
                    "CONTRATO DE TRABALHO": "CRF",
                    "FICHA DE ENTREGA DE EPI": "EPI",
                    "CERTIFICADO NR 10": "NR10",
                    "CERTIFICADO NR 11": "NR11",
                    "CERTIFICADO NR 12": "NR12",
                    "CERTIFICADO NR 33": "NR33",
                    "CERTIFICADO NR 35": "NR35",
                    "CERTIFICADO OU REGISTRO DE CLASSE SUPERIOR E/OU TÉCNICO": "DIPLOMA"
                })

        return tabela_sertras 

    @staticmethod
    def ajustar_largura_colunas(ws):
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter 
            max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
            ws.column_dimensions[col_letter].width = max_length + 2

    def personalizar_excel(self, caminho_saida):
        wb = load_workbook(caminho_saida)

        alinhamento_central = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fundo_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        fonte_branca = Font(color="FFFFFF", bold=True, size=12)

        borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for ws in wb.worksheets:  
            ws.row_dimensions[1].height = 30

            for cell in ws[1]:  
                if isinstance(cell, MergedCell):  
                    continue
                cell.fill = fundo_preto
                cell.font = fonte_branca
                cell.alignment = alinhamento_central
                cell.border = borda

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                ws.row_dimensions[row[0].row].height = 23
                for cell in row:
                    if isinstance(cell, MergedCell):  
                        continue
                    cell.border = borda 
                    cell.alignment = alinhamento_central

            self.ajustar_largura_colunas(ws)
            ws.freeze_panes = "B2"
            ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).coordinate}"

        wb.save(caminho_saida)

    def criar_excel(self,tabela_pessoas, tabela_empresa):   
        data_atual = datetime.now().strftime("%d-%m-%Y")
        caminho_saida = f"RELATÓRIO_SERTRAS {self.contrato_selecionado} {data_atual}.xlsx"

        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            tabela_pessoas.to_excel(writer, sheet_name='PESSOAS', index=False)
            tabela_empresa.to_excel(writer, sheet_name='EMPRESA', index=False)

        self.personalizar_excel(caminho_saida)

    def BaixarRelatório(self):
        self.driver = self.initialize_driver()
        doc_dp, doc_qsms, dir_dp, dir_qsms, mapeamento_para_documentos, mapeamento_para_datas, mapeamento_para_comentarios, xpath_botão_contrato = self.get_info_contrato()
        self.login_sertras(xpath_botão_contrato)

        diretorio_downloads = os.path.expanduser("~/Downloads")

        # Baixar RELATÓRIO DE EMPRESAS
        self.download_arquivo(tipo="empresas")
        caminho_empresa = self.wait_for_download(diretorio_downloads)
        tabela_empresa = self.ler_xml(caminho_empresa)
        tabela_empresa = self.tratar_tabela(tabela_empresa)

        time.sleep(1)

        # Baixar RELATÓRIO DE PESSOAS
        self.download_arquivo(tipo="pessoas")
        caminho_pessoas = self.wait_for_download(diretorio_downloads)
        tabela_pessoas = self.ler_xml(caminho_pessoas)
        tabela_pessoas = self.tratar_tabela(tabela_pessoas)

        self.criar_excel(tabela_pessoas, tabela_empresa)   

    def GerarRelatório(self):
        data_atual = datetime.now().strftime("%d-%m-%Y")
        caminho_saida = (f"RELATÓRIO_SERTRAS {self.contrato_selecionado} {data_atual}.xlsx")

        self.BaixarRelatório()

        subprocess.run(["cmd", "/c", "start", "", caminho_saida], shell=True)

# -----------------------------|------------------------------|---------------------------|-----

    @staticmethod
    def calcular_vencimento(data_str, anos=1):
        data_obj = datetime.strptime(data_str, "%d/%m/%Y")
        return (data_obj.replace(year=data_obj.year + anos)).strftime("%d/%m/%Y")

    def ler_aso(self,caminho_arquivo, poppler_path):
        paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path)
        texto_extraido = ""

        for pagina_imagem in paginas_imagem:
            texto_extraido += pytesseract.image_to_string(pagina_imagem)

        padrao_data = r'\b\d{2}/\d{2}/\d{4}\b'
        datas = re.findall(padrao_data, texto_extraido)

        if len(datas) > 1:
            return datas[-1], self.calcular_vencimento(datas[-1])  
        return None, None  

    def ler_epi(self,caminho_arquivo, poppler_path):
        paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path)
        texto_extraido = ""

        for pagina_imagem in paginas_imagem: 
            pagina = pagina_imagem.convert('L')  # escala de cinza
            pagina = ImageOps.autocontrast(pagina)
            pagina = pagina.point(lambda x: 0 if x < 128 else 255, '1')  # binarizar

            if pagina.mode not in ("RGB", "L"):
                pagina = pagina.convert("RGB")

            pagina = ImageEnhance.Sharpness(pagina).enhance(2.0)
            texto_extraido += pytesseract.image_to_string(pagina)

        padrao_data = r'\b\d{2}/\d{2}/\d{2}\b'
        datas = re.findall(padrao_data, texto_extraido)

        if datas:
            data = datas[-1]
            if len(data.split('/')[2]) == 2:
                data = data[:6] + '20' + data[6:]
            return data, self.calcular_vencimento(data)
        return None, None

    def ler_Nrs(self,caminho_arquivo, poppler_path, documento):
        paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path, dpi=300)

        texto_extraido = ""
        pagina = paginas_imagem[0]

        try:
            texto_orientacao = pytesseract.image_to_osd(pagina)
            rotacao = int(re.search(r'Rotate: (\d+)', texto_orientacao).group(1))
            if rotacao != 0:
                pagina = pagina.rotate(-rotacao, expand=True)
        except pytesseract.TesseractError as e:
            print(f"Erro ao detectar orientação: {e}", caminho_arquivo)
       
        if pagina.mode != 'L':
            pagina = pagina.convert('L')

        pagina = ImageOps.autocontrast(pagina)
        pagina = pagina.point(lambda x: 0 if x < 128 else 255, '1')

        if pagina.mode != "RGB":
            pagina = pagina.convert("RGB")

        pagina = ImageEnhance.Sharpness(pagina).enhance(2.0)

        try:
            texto_extraido += pytesseract.image_to_string(pagina)

        except pytesseract.TesseractError as e:
            print(f"Erro ao extrair texto: {e}", caminho_arquivo)
            return None, None

        padrao_data = r'(\d{1,2}\/\d{1,2}\/\d{4})|(\d{1,2}\sde\s[a-zà-ú]+\sde\s\d{4})'
        datas_encontradas = re.findall(padrao_data, texto_extraido, re.IGNORECASE)

        meses = {
            'janeiro': '01', 'fevereiro': '02', 'março': '03', 'abril': '04',
            'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08',
            'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12'
        }

        if datas_encontradas:
            for data in datas_encontradas:
                data = data[0] or data[1]
                data = data.strip().lower()
                data = re.sub(r'\s+', ' ', data)

                # Se a data estiver no formato "DD de mês de YYYY"
                if 'de' in data:
                    partes = data.split(' de ')
                    if len(partes) == 3:
                        dia = partes[0].zfill(2)
                        mes = meses.get(partes[1])
                        if mes:
                            ano = partes[2]
                            data = f"{dia}/{mes}/{ano}"

                # Define o vencimento baseado no tipo de NR
                anos = 2 if documento in ["NR35", "NR10"] else 1
                return data, self.calcular_vencimento(data, anos)

        return None, None
    
    def extrair_vencimento(self,caminho_arquivo, poppler_path, documento):
        if documento == "ASO":
            return self.ler_aso(caminho_arquivo, poppler_path)
        
        elif documento == "EPI":
            return self.ler_epi(caminho_arquivo, poppler_path)
        
        else:
            return self.ler_Nrs(caminho_arquivo, poppler_path, documento)

    @staticmethod
    def obter_data_modificacao(caminho_arquivo):
        return datetime.fromtimestamp(os.path.getmtime(caminho_arquivo))

    def verificar_atualizacao(self,status, data_analise, data_envio, caminho_arquivo):
        data_modificacao = self.obter_data_modificacao(caminho_arquivo)

        if status == "Pendente Correção":
            data_analise = datetime.strptime(data_analise, "%d/%m/%Y")
            limite_correcao = data_analise + timedelta(days=1)
            return data_modificacao > limite_correcao, data_modificacao.strftime("%d/%m/%Y %H:%M")
        
        else:
            data_envio = datetime.strptime(data_envio, "%d/%m/%Y %H:%M")
            limite_correcao = data_envio + timedelta(days=90)
            return data_modificacao > limite_correcao, data_modificacao.strftime("%d/%m/%Y %H:%M")
        
    def get_dados(self):
        data_atual = datetime.now().strftime("%d-%m-%Y")
        caminho_excel = f"RELATÓRIO_SERTRAS {self.contrato_selecionado} {data_atual}.xlsx"

        if not os.path.exists(caminho_excel):
            self.BaixarRelatório() 
            tabela_sertras = pd.read_excel(caminho_excel, sheet_name="PESSOAS")

        else:
            tabela_sertras = pd.read_excel(caminho_excel)

        return tabela_sertras

    def interacao_interface_recursos(self):
        botão_recursos = WebDriverWait(self.driver,10).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="sidebar-menu"]/div/ul/li[8]/a/span[1]')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_recursos)
        botão_recursos.click()

        botão_recursos_pessoas = WebDriverWait(self.driver,10).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="sidebar-menu"]/div/ul/li[8]/ul/li[1]/a')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_recursos_pessoas)
        botão_recursos_pessoas.click()

    def interacao_interface_envio(self, nome):
        campo_nome = WebDriverWait(self.driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="filtro_nome"]'))) 
        campo_nome.clear()
        campo_nome.send_keys(nome)

        botão_filtrar_nome = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH,  '//*[@id="dashboard-v1"]/div[4]/div/div/div[2]/form/div[6]/button[1]')))
        botão_filtrar_nome.click()

        for tentativa in range(4):
            try:
                botao_eventos = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="data-tables2"]/tbody/tr/td[9]/a')))
                botao_eventos.click()
                break

            except StaleElementReferenceException:
                print(f"Tentativa {tentativa+1}/5 falhou: Botão de eventos foi recriado. Tentando novamente...")

        abas = self.driver.window_handles
        self.driver.switch_to.window(abas[-1])

        # Bloco para garantir o envio de documentos de funcionários demitidos, visto que o xpath do botão de documentação desses são alterados
        try:   
            botão_documentação = WebDriverWait(self.driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="data-tables2"]/tbody/tr[4]/td[4]/ul/li/a')))

        except TimeoutException:
            try:
                botão_documentação = WebDriverWait(self.driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="data-tables2"]/tbody/tr[5]/td[4]/ul/li/a')))

            except TimeoutException:
                print(f"Nenhum botão de documentação encontrado para {nome}")
                st.error(f"Nenhum botão de documentação encontrado para {nome}")
                return

        self.driver.execute_script("arguments[0].scrollIntoView();", botão_documentação)
        botão_documentação.click()

        abas = self.driver.window_handles
        self.driver.switch_to.window(abas[-1])

    def enviar_documento(self, documentos_validos, mapeamento_para_documentos, mapeamento_para_datas, mapeamento_para_comentarios,vencimentos_enviados, documentos_enviados):
        for status, arquivo, documento, caminho_arquivo, data_vencimento, função in documentos_validos:
            if documento in mapeamento_para_datas:
                xpath_data = mapeamento_para_datas[documento]
                campo_data = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_data)))
                self.driver.execute_script("arguments[0].scrollIntoView();", campo_data)

                if status == "Pendente Correção":
                    data_vencimento = campo_data.get_attribute("value")
                    data_vencimento = datetime.strptime(data_vencimento, "%d/%m/%Y")
                    data_vencimento += timedelta(days=1)
                    data_vencimento = data_vencimento.strftime('%d/%m/%Y')

                campo_data.clear()
                campo_data.send_keys(data_vencimento)
                vencimentos_enviados.append(data_vencimento)

            if documento in mapeamento_para_documentos:
                xpath_documento = mapeamento_para_documentos[documento]
                botao_upload = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_documento)))
                self.driver.execute_script("arguments[0].scrollIntoView();", botao_upload)
                botao_upload.send_keys(caminho_arquivo)
                documentos_enviados.append(arquivo)

            if documento in mapeamento_para_comentarios:
                xpath_comentario = mapeamento_para_comentarios[documento]
                campo_comentario = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_comentario)))
                self.driver.execute_script("arguments[0].scrollIntoView();", campo_comentario)
                campo_comentario.clear()

            if data_vencimento is None:
                vencimentos_enviados.append("N/A")

            time.sleep(1)

        botao_envio = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnFuncaoRequisitoValores"]')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botao_envio)
        botao_envio.click()

        time.sleep(2.5)

        abas = self.driver.window_handles
        self.driver.switch_to.window(abas[-1])
        self.driver.close()
        self.driver.switch_to.window(abas[-2])
        self.driver.close()
        self.driver.switch_to.window(abas[0])

    def EnvioSertras(self):
        doc_dp, doc_qsms, dir_dp, dir_qsms, mapeamento_para_documentos, mapeamento_para_datas, mapeamento_para_comentarios, xpath_botão_contrato = self.get_info_contrato()

        tabela_sertras = self.get_dados()
        tabela_sertras = tabela_sertras.drop(["COMENTÁRIO ANALISTA", "PRAZO SLA"], axis=1)

        Status = ["Pendente", "Pendente Correção", "Vencido"]
        tabela_sertras = tabela_sertras[tabela_sertras["STATUS"].isin(Status)]

        try:
            if not hasattr(self, 'driver') or self.driver is None:
                self.driver = self.initialize_driver()
                self.login_sertras(xpath_botão_contrato)

            else:
                titulo = self.driver.title
                if "Sertras - Sistema de Gestão de Terceiros - Fornecedores" not in titulo: 
                    raise Exception("Não está na plataforma correta!")
                
        except Exception as e:
            self.driver = self.initialize_driver()
            self.login_sertras(xpath_botão_contrato)

        self.interacao_interface_recursos()

        documentos_enviados = []
        erro_envio = []
        documentos_não_encontrados = []
        documentos_encontrados = []
        documentos_atualizados = []
        documentos_nao_atualizados = []
        datas_extraidas = []
        datas_modificacao = []
        vencimentos_projetados = []
        vencimentos_enviados = []

        for nome, grupo in tabela_sertras.groupby("NOME"):
            documentos_validos = []

            for _, linha in grupo.iterrows():
                status, documento, funcao = linha["STATUS"], linha["DOCUMENTO"], linha["FUNÇÃO"]
                caminho_base = os.path.join(os.path.expanduser('~'), *dir_dp) if documento in doc_dp else os.path.join(os.path.expanduser('~'), *dir_qsms)
                arquivo = f"{documento} - {nome}"
                
                caminho_arquivo = os.path.join(os.path.expanduser("~"), caminho_base, nome, f"{arquivo}.pdf")

                if not os.path.exists(caminho_arquivo):
                    documentos_não_encontrados.append(arquivo)
                    continue

                documentos_encontrados.append(arquivo)

                if status in ["Pendente Correção","Vencido"]:
                    atualizado, data_modificacao = self.verificar_atualizacao(status, linha["DATA ANÁLISE"], linha["DATA ENVIO"], caminho_arquivo)
                    datas_modificacao.append(data_modificacao)

                    if not atualizado:
                        documentos_nao_atualizados.append(arquivo)
                        continue
                    
                    else:
                        documentos_atualizados.append(arquivo)

                else:
                    datas_modificacao.append("N/A")

                if documento in ["ASO","EPI", "NR10", "NR11", "NR12", "NR33", "NR35"]:
                    if status in ["Pendente", "Vencido"]:
                        data_extraida, data_vencimento = self.extrair_vencimento(caminho_arquivo, poppler_path, documento)

                        if isinstance(data_vencimento, (list, tuple)):
                            data_vencimento = data_vencimento[0] if data_vencimento else None

                        if not data_vencimento:
                            erro_envio.append(arquivo)
                            continue

                        try:
                            data_vencimento = datetime.strptime(data_vencimento, "%d/%m/%Y")

                            if status == "Pendente Correção":
                                data_vencimento += timedelta(days=1)
                            
                            data_vencimento = data_vencimento.strftime('%d/%m/%Y')

                        except (ValueError, TypeError):
                            erro_envio.append(arquivo)
                            continue

                        datas_extraidas.append(data_extraida)
                        vencimentos_projetados.append(data_vencimento)

                    else:
                        data_vencimento = None
                        datas_extraidas.append("N/A")
                        vencimentos_projetados.append("N/A")

                else:
                    data_vencimento = None
                    datas_extraidas.append("N/A")
                    vencimentos_projetados.append("N/A")

                documentos_validos.append((status, arquivo, documento, caminho_arquivo, data_vencimento, funcao))

            if documentos_validos:
                self.interacao_interface_envio(nome)
            
                self.enviar_documento(documentos_validos, 
                                mapeamento_para_documentos.get(funcao, mapeamento_para_documentos["OUTRAS"]), 
                                mapeamento_para_datas.get(funcao, mapeamento_para_datas["OUTRAS"]),
                                mapeamento_para_comentarios.get(funcao, mapeamento_para_comentarios["OUTRAS"]), 
                                vencimentos_enviados, documentos_enviados)
    
        self.driver.quit()

        return tabela_sertras, documentos_não_encontrados, documentos_encontrados, documentos_enviados, datas_extraidas, vencimentos_projetados, vencimentos_enviados, erro_envio, documentos_atualizados, documentos_nao_atualizados, datas_modificacao 


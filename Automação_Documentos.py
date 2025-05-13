import os
import re 
import subprocess
import locale
import shutil
import zipfile
import pytesseract
import pandas as pd
from dateutil.relativedelta import relativedelta
from docx import Document
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from PIL import Image as PILImage
from pdf2image import convert_from_path
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.views import SheetView
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
 

locale.setlocale(locale.LC_TIME, "pt_BR.utf8")
PILImage.MAX_IMAGE_PIXELS = None

poppler_path = os.path.join(os.path.expanduser("~"),"Downloads","Release-24.08.0-0","poppler-24.08.0","Library","bin")
tesseract_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = tesseract_path

class AutomaçãoDocumentos():

    def __init__(self, contratos, contrato_selecionado):
        self.contratos = contratos
        self.poppler_path = poppler_path
        self.contrato_selecionado = contrato_selecionado

    def get_info_contrato(self, chamado=None):
        path_efetivo = self.contratos[self.contrato_selecionado]["diretorio efetivo"]
        path_funcionários = self.contratos[self.contrato_selecionado]["diretorio funcionarios"]["QSMS"]
        path_modelos = self.contratos[self.contrato_selecionado]["diretorio modelos"]
        path_saídas = self.contratos[self.contrato_selecionado]["diretorio saida"]
        documentos_por_função = self.contratos[self.contrato_selecionado]["documentos/função"]

        diretorio_efetivo = os.path.join(os.path.expanduser("~"), *path_efetivo)
        diretorio_funcionarios = os.path.join(os.path.expanduser("~"), *path_funcionários)
        diretorio_modelos = os.path.join(os.path.expanduser("~"), *path_modelos)
        diretorio_saidas = os.path.join(os.path.expanduser("~"), *path_saídas)

        if chamado == 'RelatórioDocumentos':
            return diretorio_efetivo, diretorio_funcionarios, documentos_por_função
        
        if chamado == "GerarDocumentos":
            return diretorio_efetivo, diretorio_funcionarios, documentos_por_função, diretorio_modelos, diretorio_saidas
    
    def ler_aso(self,caminho_arquivo, poppler_path):
        paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path)
        texto_extraido = ""

        for pagina_imagem in paginas_imagem:
            texto_extraido += pytesseract.image_to_string(pagina_imagem)

        padrao_data = r'\b\d{2}/\d{2}/\d{4}\b'
        datas = re.findall(padrao_data, texto_extraido)

        if len(datas) > 1:
            return datas[-1] 
        
        return None  

    def ler_epi(self,caminho_arquivo, poppler_path):
        paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path)
        texto_extraido = ""

        for pagina_imagem in paginas_imagem:
            texto_extraido += pytesseract.image_to_string(pagina_imagem)

        padrao_data = r'\b\d{2}/\d{2}/\d{2}\b'
        datas = re.findall(padrao_data, texto_extraido)

        if datas:
            data = datas[-1]
            if len(data.split('/')[2]) == 2:
                data = data[:6] + '20' + data[6:]

            return data
        
        return None

    def ler_Nrs(self,caminho_arquivo, poppler_path):
        paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path, dpi=300)
        texto_extraido = ""

        for pagina in paginas_imagem:
            texto_orientacao = pytesseract.image_to_osd(pagina)
            rotacao = int(re.search(r'Rotate: (\d+)', texto_orientacao).group(1))
            if rotacao != 0:
                pagina = pagina.rotate(-rotacao, expand=True)
            
            texto_extraido += pytesseract.image_to_string(pagina)

        padrao_data = r'(\d{1,2}\/\d{1,2}\/\d{4})|(\d{1,2}\sde\s[a-zà-ú]+\sde\s\d{4})'
        datas_encontradas = re.findall(padrao_data, texto_extraido, re.IGNORECASE)

        meses = {
            'janeiro': '01', 'fevereiro': '02', 'março': '03', 'marco': '03','abril': '04',
            'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08',
            'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12'
        }

        if datas_encontradas:
            for data in datas_encontradas:
                data = data[0] or data[1]
                data = data.strip().lower()
                data = re.sub(r'\s+', ' ', data)

                if 'de' in data:
                    partes = data.split(' de ')
                    if len(partes) == 3:
                        dia = partes[0].zfill(2)
                        mes = meses.get(partes[1])
                        if mes:
                            ano = partes[2]
                            data = f"{dia}/{mes}/{ano}"

                return data

        return None
    
    def ler_OS(self,caminho_arquivo, poppler_path):
        paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path)
        texto_extraido = ""

        for pagina_imagem in paginas_imagem:
            texto_extraido += pytesseract.image_to_string(pagina_imagem)

        padrao_data = r'\b\d{2}/\d{2}/\d{2,4}\b'
        datas = re.findall(padrao_data, texto_extraido)

        if datas:
            data = str(datas[-1])
            dia, mes, ano = data.split("/")

            if len(ano) == 2:
                ano = "20" + ano
                data = f"{dia}/{mes}/{ano}"

            return data
        
        return None
    
    def extrair_data(self,caminho_arquivo, poppler_path, documento):
        if documento == "ASO":
            return self.ler_aso(caminho_arquivo, poppler_path)
        
        elif documento == "EPI":
            return self.ler_epi(caminho_arquivo, poppler_path)
        
        elif documento == "OS":
            return self.ler_OS(caminho_arquivo, poppler_path)
        
        elif documento in ["NR6", "NR10", "NR11", "NR12", "NR18", "NR33", "NR35"]:
            return self.ler_Nrs(caminho_arquivo, poppler_path)
        
        else:
            return self.ler_aso(caminho_arquivo, poppler_path)
    
    @staticmethod
    def get_diretorio_funcionario(diretorio_funcionarios, nome):
        return os.path.join(os.path.expanduser("~"), diretorio_funcionarios, nome)

    @staticmethod
    def formatar_cpf(cpf):
        cpf = ''.join(filter(str.isdigit, str(cpf)))
        cpf = cpf.zfill(11)
        return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
    
    @staticmethod
    def formatar_data(data):
        return datetime.strftime(data, '%d/%m/%Y')
    
    @staticmethod
    def obter_documentos_requeridos(funcao, documentos_por_funcao):
        return documentos_por_funcao.get(funcao, documentos_por_funcao["OUTRAS"])

    @staticmethod
    def tratar_tabela(tabela_dados):
        tabela_dados["DESC FUNÇÃO"] = tabela_dados["DESC FUNÇÃO"].replace({"1/2 OFICIAL DE REPARO DE REDE DE SANEAMENTO CIVIL" : "MEIO OFICIAL DE REPARO DE REDE DE SANEAMENTO CIVIL",
                        "ASSISTENTE PLANEJAMENTO II" : "ASSISTENTE PLANEJAMENTO",
                        "ASSISTENTE PLANEJAMENTO III" : "ASSISTENTE PLANEJAMENTO",
                        "AUXILIAR ADMINISTRATIVO I" : "AUXILIAR ADMINISTRATIVO",
                        "AUXILIAR ADMINISTRATIVO III" : "AUXILIAR ADMINISTRATIVO",
                        "ELETRICISTA DE REPARO DE REDE DE SANEAMENTO I" : "ELETRICISTA DE REPARO DE REDE DE SANEAMENTO",
                        "ENCARREGADO DE REPARO DE REDE DE SANEAMENTO I" : "ENCARREGADO DE REPARO DE REDE DE SANEAMENTO",
                        "ENCARREGADO DE REPARO DE REDE DE SANEAMENTO II" : "ENCARREGADO DE REPARO DE REDE DE SANEAMENTO",
                        "TECNICO SEGURANCA DO TRABALHO PL" : "TECNICO SEGURANCA DO TRABALHO",
                        })
        
        return tabela_dados
    
    @staticmethod
    def ajustar_largura_colunas(ws):
        tamanho_minimo_documentos_pendentes = 28
        tamanho_minimo_documentos = 11
        tamanho_minimo_nome = 25

        for idx, column_cells in enumerate(ws.columns):
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            valid_cells = [cell for cell in column_cells if not isinstance(cell, MergedCell) and cell.value]
            if not valid_cells:
                continue
            column = valid_cells[0].column_letter

            if idx == 0:  
                ws.column_dimensions[column].width = max(max_length + 2, tamanho_minimo_nome)

            elif idx >= 4:  
                ws.column_dimensions[column].width = tamanho_minimo_documentos

            else:
                ws.column_dimensions[column].width = max_length + 2

    def personalizar_planilha(self, ws):
        fundo_azul = PatternFill(start_color="003399", end_color="003399", fill_type="solid")
        fonte_branca = Font(size=12, color="FFFFFF", bold=True)
        verde_claro = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  
        vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        laranja = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")

        alinhamento_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)
        alinhamento_central = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)

        borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        ws.sheet_view.view = "pageBreakPreview"

        img = ExcelImage(r"img\LOGO CONCREJATO.png")  
        img.width = 250
        img.height = 70
        ws.add_image(img, "A1")

        ws["A1"] = "CONTROLE DE DOCUMENTAÇÃO FUNCIONÁRIOS"
        ws["A1"].font = Font(bold=True, size=30, color="003399")
        ws["A1"].alignment = alinhamento_central

        ws.row_dimensions[1].height = 80
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column) 

        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 28

        for cell in ws[2]:
            cell.fill = fundo_azul
            cell.font = fonte_branca
            cell.alignment = alinhamento_central
            cell.border = borda

        for cell in ws[3]:
            cell.fill = fundo_azul
            cell.font = fonte_branca
            cell.alignment = alinhamento_central
            cell.border = borda

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 25
            for idx, cell in enumerate(row):
                cell.border = borda
                if idx == 0:  
                    cell.alignment = alinhamento_esquerda

                else:  
                    cell.alignment = alinhamento_central

                if cell.value == "OK":
                    cell.fill = verde_claro

                elif cell.value == "P":
                    cell.fill = vermelho

                elif cell.value == "E":
                    cell.fill = laranja

        self.ajustar_largura_colunas(ws)

        ws.freeze_panes = "D4"
        ws.auto_filter.ref = "A3:O{}".format(ws.max_row)

    def gerar_dados_planilha(self, diretorio_funcionarios, tabela_dados, documentos_por_função, fixos, mesclados, ws):
        dados_planilha = []

        col = 1
        for item in fixos:
            ws.cell(row=2, column=col, value=item)
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            col += 1

        for exig in mesclados:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
            ws.cell(row=2, column=col, value=exig)

            ws.cell(row=3, column=col, value="STATUS").alignment = Alignment(horizontal="center")
            ws.cell(row=3, column=col+1, value="DATA").alignment = Alignment(horizontal="center")
            col += 2

        linha_planilha = 4  
        for funcionário, linha in tabela_dados.groupby("NOME"):
            caminho_funcionario = os.path.join(os.path.expanduser('~'), diretorio_funcionarios, funcionário)
            funcao = str(linha["DESC FUNÇÃO"].iloc[0])
            admissao = str(linha["DATA ADMISSAO"].iloc[0])
            admissao = pd.to_datetime(admissao).strftime("%d/%m/%Y")
            cpf = str(linha["CPF"].iloc[0])
            cpf = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"  

            ws.cell(row=linha_planilha, column=1, value=funcionário)
            ws.cell(row=linha_planilha, column=2, value=funcao)
            ws.cell(row=linha_planilha, column=3, value=cpf)
            ws.cell(row=linha_planilha, column=4, value=admissao)

            documentos_requeridos = self.obter_documentos_requeridos(funcao, documentos_por_função)

            col_exig = 5  
            for doc in mesclados:
                if doc in documentos_requeridos:
                    arquivo = f"{doc} - {funcionário}.pdf"
                    caminho_arquivo = os.path.join(caminho_funcionario, arquivo)
                
                    if os.path.exists(caminho_arquivo):
                        status = "OK"
                        print(f"Extraindo a data do arquivo: {arquivo}")    
                        data = self.extrair_data(caminho_arquivo, self.poppler_path, doc)

                        if data is None:
                            data = 'E'

                    else:
                        status = "P"
                        data = "P"

                else:
                    status = "NA"
                    data = "NA"

                ws.cell(row=linha_planilha, column=col_exig, value=status)
                ws.cell(row=linha_planilha, column=col_exig+1, value=data)
                col_exig += 2

            dados_planilha.append(linha_planilha)
            linha_planilha += 1 

        self.personalizar_planilha(ws)
        
        return dados_planilha

    def GerarRelatório(self):
        wb = Workbook()
        fixos = ["FUNCIONÁRIO", "FUNÇÃO", "CPF", "ADMISSÃO"]
        mesclados = ["FRE", "ASO", "EPI", "NR6", "NR10", "NR11", "NR12", "NR18", "NR33", "NR35", "OS"]

        diretorio_efetivo, diretorio_funcionarios, documentos_por_função = self.get_info_contrato('RelatórioDocumentos')

        tabela_dados = pd.read_excel(diretorio_efetivo)
        tabela_dados = self.tratar_tabela(tabela_dados)

        ws = wb.create_sheet(title=self.contrato_selecionado) 
        dados_planilha = self.gerar_dados_planilha(diretorio_funcionarios, tabela_dados, documentos_por_função, fixos, mesclados, ws)
        
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        data_atual = datetime.now().strftime("%d-%m-%Y")
        caminho_saida =  f"RELATÓRIO_DOCUMENTAÇÃO {self.contrato_selecionado} {data_atual}.xlsx"
        wb.save(caminho_saida)

        return caminho_saida

    def ExibirRelatório(self):
        caminho_saida = self.GerarRelatório()
        subprocess.run(["cmd", "/c", "start", "", caminho_saida], shell=True)


# --------------------------|----------------------------------|-----------------------------------------------|---------------------------------
        
    def get_modelo(self, documento, funcao, contrato, diretorio_modelos):      
        if not diretorio_modelos:
            print(f"[ERRO] Diretório de modelos não encontrado para o contrato: {contrato}")
            return None
        
        mapa_modelos = {
            doc: os.path.join(diretorio_modelos, "NRs - MODELOS", f"{doc} - MODELO.docx")
            for doc in ["NR6", "NR12", "NR18", "NR33", "NR35"]
        }
        mapa_modelos["OS"] = os.path.join(diretorio_modelos, "OS - MODELOS", f"OS - {funcao}.docx")
        
        modelo = mapa_modelos.get(documento)
        
        if not modelo or not os.path.exists(modelo):
            print(f"[ERRO] Modelo não encontrado ou inexistente para {documento} ({funcao}) no contrato {contrato}.")
            return None
        
        return modelo

    @staticmethod
    def substituir_texto_docx(nome_modelo, substituicoes, diretorio_saida):
        temp_zip_path = diretorio_saida.replace(".docx", "_temp.zip")
        temp_folder = diretorio_saida.replace(".docx", "_temp")
        
        shutil.copy2(nome_modelo, diretorio_saida)
        with zipfile.ZipFile(diretorio_saida, 'r') as docx_zip:
            docx_zip.extractall(temp_folder)
        
        xml_path = os.path.join(temp_folder, "word", "document.xml")
        with open(xml_path, "r", encoding="utf-8") as file:
            xml_content = file.read()
        
        for marcador, novo_texto in substituicoes.items():
            xml_content = re.sub(re.escape(marcador) + r"(\s*</w:t>\s*<w:t[^>]*>)?", novo_texto, xml_content)
        
        with open(xml_path, "w", encoding="utf-8") as file:
            file.write(xml_content)
        
        with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as docx_zip:
            for root, _, files in os.walk(temp_folder):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_folder)
                    docx_zip.write(file_path, arcname)
        
        os.replace(temp_zip_path, diretorio_saida)
        shutil.rmtree(temp_folder, ignore_errors=True)

    def gerar_documentos_pendentes(self, nome_funcionario, funcao, cpf, admissao, documentos_pendentes, diretorio_modelos, diretorio_saidas):
        for documento in documentos_pendentes:
            modelo = self.get_modelo(documento, funcao, self.contrato_selecionado, diretorio_modelos)

            if not modelo:
                print(f"[AVISO] Documento {documento} não criado para {nome_funcionario} ({funcao}) no contrato {self.contrato_selecionado}")
                continue  
            
            caminho_saida = f"{diretorio_saidas}/{documento}/{documento} - {nome_funcionario}.docx"
            admissao_formatada = datetime.strptime(admissao, "%d/%m/%Y").strftime("%d de %B de %Y") if documento.startswith("NR") else admissao
            
            substituicoes = {
                "{{NOME}}": nome_funcionario,
                "{{FUNÇÃO}}": funcao,
                "{{CPF}}": cpf,
                "{{ADMISSÃO}}": admissao_formatada,
                "{{TREINAMENTO}}": admissao_formatada
            }
            
            self.substituir_texto_docx(modelo, substituicoes, caminho_saida)

    def gerar_dados(self,diretorio_funcionarios, tabela_dados, documentos_por_função, ordem_documentos):
        dados_planilha = []
        cabeçalho = ["FUNCIONÁRIO", "FUNÇÃO", "CPF", "ADMISSÃO", "FRE", "ASO", "EPI", "NR6", "NR10", "NR11", "NR12", "NR18", "NR33", "NR35", "OS", "DOCUMENTOS PENDENTES"]
        
        for nome, linha in tabela_dados.groupby("NOME"):
            caminho_funcionario = self.get_diretorio_funcionario(diretorio_funcionarios, nome)
            funcao = str(linha["DESC FUNÇÃO"].iloc[0])
            admissao = self.formatar_data(linha["DATA ADMISSAO"].iloc[0])
            cpf = self.formatar_cpf(linha["CPF"].iloc[0])

            documentos_requeridos = self.obter_documentos_requeridos(funcao, documentos_por_função)
            documentos_na_pasta = os.listdir(caminho_funcionario) if os.path.isdir(caminho_funcionario) else []
            
            linha_dados = [nome, funcao, cpf, admissao]
            documentos_pendentes = []

            for documento in ordem_documentos:
                if documento in documentos_requeridos:
                    nome_esperado = f"{documento} - {nome}.pdf"
                    linha_dados.append("OK" if nome_esperado in documentos_na_pasta else "P")

                    if nome_esperado not in documentos_na_pasta:
                        documentos_pendentes.append(documento)
                else:
                    linha_dados.append("N/A")
            
            linha_dados.append(" - ".join(documentos_pendentes) if documentos_pendentes else "---")
            dados_planilha.append(linha_dados)

        return pd.DataFrame(dados_planilha, columns=cabeçalho)
    
    def GerarDocumentos(self):
        data_atual = datetime.now().strftime("%d-%m-%Y")
        diretorio_efetivo, diretorio_funcionarios, documentos_por_função, diretorio_modelos, diretorio_saidas = self.get_info_contrato('GerarDocumentos')

        ordem_documentos = ["FRE", "ASO", "EPI", "NR6", "NR10", "NR11", "NR12", "NR18", "NR33", "NR35", "OS"]

        tabela_dados = pd.read_excel(diretorio_efetivo)
        tabela_dados = self.tratar_tabela(tabela_dados)

        tabela = self.gerar_dados(diretorio_funcionarios, tabela_dados, documentos_por_função, ordem_documentos)
            
        for _, row in tabela.iterrows():
            nome_funcionario = row["FUNCIONÁRIO"]
            funcao = row["FUNÇÃO"]
            cpf = row["CPF"]
            admissao = row["ADMISSÃO"]
            documentos_pendentes = [doc for doc in tabela.columns[4:] if row[doc] == "P"]
            
            if documentos_pendentes:
                self.gerar_documentos_pendentes(nome_funcionario, funcao, cpf, admissao, documentos_pendentes,diretorio_modelos, diretorio_saidas)

